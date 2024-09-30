import os,math
import cv2,numpy
import shutil
import re
from mako.template import Template
from ruamel import yaml
from ..models.action import SwipeAction, PressKeyboardAction, UnknownAction
from ..constants import *


def write_image(path, image):
    name, ext = os.path.splitext(path)
    cv2.imencode(ext, image)[1].tofile(path)
    # Not work for non-English path
    #cv2.imwrite(path, image)


def export_previews(actions, root_export_dir):
    preview_export_dir = os.path.join(root_export_dir, "previews")
    if os.path.exists(preview_export_dir):
        raise Exception("Preview directory already exist. Please first remove it!")
    os.makedirs(preview_export_dir)
    for action in actions:
        index = action.index
        preview_image = action.get_preview_image()
        preview_name = "preview%d.png"%index
        preview_export_path = os.path.join(preview_export_dir, preview_name)
        write_image(preview_export_path, preview_image)

def export_pre_action_frames(actions, root_export_dir):
    frame_export_dir = os.path.join(root_export_dir, "frames")
    if os.path.exists(frame_export_dir):
        raise Exception("Frame directory already exist. Please first remove it!")
    os.makedirs(frame_export_dir)
    for action in actions:
        index = action.index
        device_image = action.pre_action_frame.get_device_image()
        frame_name = "frame%d.png"%index
        frame_export_path = os.path.join(frame_export_dir, frame_name)
        write_image(frame_export_path, device_image)

def generate_script(actions, root_export_dir):
    if os.path.exists(root_export_dir):
        files = os.listdir(root_export_dir)
        if "debug" in files:
            files.remove("debug")
        if len(files) > 0:
            raise Exception("Script folder %s already exist. Please remove it before recording!"% root_export_dir)
    else:
        os.makedirs(root_export_dir)

    image_export_dir = os.path.join(root_export_dir, "images")
    os.makedirs(image_export_dir)
    script_export_path = os.path.join(root_export_dir, "script.py")
    instructions = []
    for i in range(len(actions)):
        action = actions[i]
        sleep_time = None
        if i > 0:
            fps = action.parent.parameters["fps"]
            pre_action = actions[i-1]
            if len(pre_action.touch_fingertip_groups)>0:
                sleep_frame_count = action.action_fingertip_group[0].index - pre_action.touch_fingertip_groups[-1][-1].index
                sleep_time = math.ceil(sleep_frame_count/fps)
            else:
                sleep_time = 3
        name = action.get_instruction_name()
        multimedia_parameters = action.get_instruction_parameters()
        parameter_list = []
        image_index = 1
        for parameter in multimedia_parameters:
            if isinstance(parameter, numpy.ndarray):
                widget = parameter
                image_name = "image%d-%d.png"%(action.index, image_index)
                image_export_path = os.path.join(image_export_dir, image_name)
                write_image(image_export_path, widget)
                string_parameter = '"%s"' % image_name
                image_index += 1
            elif isinstance(parameter, list):
                string_parameter = "[" + ",".join("%0.3f" % d for d in parameter) + "]"
            else:
                string_parameter = '"%s"' % parameter
            parameter_list.append(string_parameter)
        combined_parameter = ",".join(parameter_list)
        instructions.append((name, combined_parameter, sleep_time))
    script_content = Template(SCRIPT_TEMPLATE_CONTENT).render(instructions=instructions, note=None)
    with open(script_export_path, "w") as script_file:
        script_file.write(script_content)
    generate_default_screenshot(actions[0], image_export_dir)

def generate_default_screenshot(action, export_dir):
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)
    screenshot_export_path = os.path.join(export_dir, "default.snapscreen")
    if os.path.exists(screenshot_export_path):
        raise Exception("screenshot file already exist. Please first remove it!")
    device_location = action.parent.parameters["device_location"]
    device_w = device_location.get_width()
    device_h = device_location.get_height()
    screenshot_content = "contour:\n  width: %d\n  height: %d"%(device_w, device_h)
    with open(screenshot_export_path, "w") as f:
        f.write(screenshot_content)

def generate_result_text(actions, export_dir):
    result = {"device":None, "actions":{}}
    device_location = actions[0].parent.parameters["device_location"]
    device_x1, device_y1 = device_location.lt
    device_x2, device_y2 = device_location.rb
    device_location_dict = {"left-top":str((device_x1, device_y1)), "right-bottom":str((device_x2, device_y2))}
    actions_dict = {}
    for i in range(len(actions)):
        action = actions[i]
        action_dict = {}
        action_dict["type"] = action.get_instruction_name()
        if isinstance(action, SwipeAction):
            params = action.get_instruction_parameters()
            action_dict["note"] = params[0]
            if len(params) > 1:
                action_dict["region"] = params[1]

            touch_positions = action.get_key_touch_positions()
            waypoints = []
            for touch_x, touch_y in touch_positions:
                device_touch_x, device_touch_y = touch_x - device_x1, touch_y - device_y1
                waypoints.append({"touch-position": str((device_touch_x, device_touch_y))})
            action_dict["waypoints"] = waypoints
        elif isinstance(action, PressKeyboardAction):
            action_dict["note"] = "|".join(action.get_instruction_parameters())
        elif isinstance(action, UnknownAction):
            pass
        else:
            touch_positions = action.get_key_touch_positions()
            widget_areas = action.get_widget_areas()
            widgets_dict = {}
            for j in range(len(widget_areas)):
                widget_area = widget_areas[j]
                touch_position = touch_positions[j]
                touch_x, touch_y = touch_position
                device_touch_x, device_touch_y = touch_x-device_x1, touch_y-device_y1
                widget_x1, widget_y1 = widget_area.lt
                widget_x2, widget_y2 = widget_area.rb
                device_widget_x1, device_widget_y1 = widget_x1-device_x1, widget_y1-device_y1
                device_widget_x2, device_widget_y2 = widget_x2-device_x1, widget_y2-device_y1
                widget_dict_area = {"left-top":str((device_widget_x1, device_widget_y1)), "right-bottom":str((device_widget_x2, device_widget_y2))}
                widget_dict = {"touch-position":str((device_touch_x, device_touch_y)), "area":widget_dict_area}
                widgets_dict["widget%d"%(j+1)] = widget_dict
            action_dict["widgets"] = widgets_dict
        actions_dict["action%d"%(i+1)] = action_dict
    result = {"device":device_location_dict, "actions":actions_dict}
    result_text_export_path = os.path.join(export_dir, "result.yaml")
    with open(result_text_export_path, "w", encoding="utf-8") as f:
        yaml.dump(result, f, Dumper=yaml.RoundTripDumper,default_flow_style=False,allow_unicode=True, indent=2)

def generate_fingertip_xls(actions, export_dir):
    from openpyxl import Workbook
    
    if "sidelook_touch_threshold" in actions[0].parent.parameters:
        sidelook_touch_threshold = actions[0].parent.parameters["sidelook_touch_threshold"]
    else:
        sidelook_touch_threshold = 0
    frame_total_count = actions[-1].action_fingertip_group[-1].index + 10
    overlook_fingertip_positions = [(0,0) for _ in range(frame_total_count)]
    sidelook_fingertip_positions = [(0,0) for _ in range(frame_total_count)]
    for action in actions:
        for fingertip in action.action_fingertip_group:
            index = fingertip.index
            overlook_fingertip_positions[index] = fingertip.overlook_position if fingertip.overlook_position != None else (0, 0)
            sidelook_fingertip_positions[index] = fingertip.sidelook_position if fingertip.sidelook_position != None else (0, 0)
    
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)
    xls_export_path = os.path.join(export_dir, "fingertip_datas.xlsx")
    wb = Workbook()
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet("Fingertip")
    else:
        ws = wb.get_sheet_by_name(wb.sheetnames[0])
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:C1')
    ws.merge_cells('D1:E1')
    ws.merge_cells('F1:F2')
    ws['A1'] = '#'
    ws['B1'] = 'overlook fingertip'
    ws['D1'] = 'sidelook fingertip'
    ws['F1'] = 'sidelook touch threshold'
    ws['B2'], ws['C2'] = "X", "Y"
    ws['D2'], ws['E2'] = "X", "Y"
    for i in range(frame_total_count):
        overlook_x, overlook_y = overlook_fingertip_positions[i]
        sidelook_x, sidelook_y = sidelook_fingertip_positions[i]
        row = (i, overlook_x, overlook_y, sidelook_x, sidelook_y, sidelook_touch_threshold)
        ws.append(row)
    wb.save(xls_export_path)

def generate(actions, export_dir, dev_mode):
    generate_script(actions, export_dir)
    export_pre_action_frames(actions, export_dir)
    if dev_mode:
        export_previews(actions, export_dir)
        generate_result_text(actions, export_dir)
        generate_fingertip_xls(actions, export_dir)


def create_script_html(script_dir):
    """ Create HTML view for a test script """
    script_file = None
    for f in os.listdir(script_dir):
        if f.endswith(".py"):
            script_file = f
            break

    if script_file is None:
        raise Exception("No script found under %s" % script_dir)

    script_path = os.path.join(script_dir, script_file)
    with open(script_path, "r") as f:
        code_lines = f.readlines()
    code_htmls = []
    for line_index in range(len(code_lines)):
        code_line = code_lines[line_index]
        image_relative_local_paths = []
        image_relative_local_paths.extend(re.findall(r'".*?\.png"', code_line))
        image_relative_local_paths.extend(re.findall(r"'.*?\.png'", code_line))
        for image_relative_local_path in image_relative_local_paths:
            image_remote_url = "images/" + image_relative_local_path[1:-1]
            code_line = code_line.replace(image_relative_local_path, '<img src="%s" />' % image_remote_url)
        color = "green" if code_line.strip().startswith("#") else "black"
        code_html = '<div><p>%02d:   </p><p style="color:%s;">%s</p></div>' % (line_index + 1, color, code_line)
        code_htmls.append(code_html)
    script_html_content = "".join(code_htmls)
    script_html_template = '<html><head><style type="text/css">p{{white-space: pre;display:inline}}</style></head><body><div id="script">{0}</div></body><script src="/jquery-4.3.1.js"></script><script src="/adjust_pic_size.js"></script><html>'
    html = script_html_template.format(script_html_content)

    script_html_file = os.path.join(script_dir, script_file + ".html")
    with open(script_html_file, "w") as f:
        f.write(html)

    jquery_file = os.path.join(os.path.dirname(__file__), "jquery-4.3.1.js")
    adjust_pic_file = os.path.join(os.path.dirname(__file__), "adjust_pic_size.js")
    shutil.copyfile(jquery_file, os.path.join(script_dir, "jquery-4.3.1.js"))
    shutil.copyfile(adjust_pic_file, os.path.join(script_dir, "adjust_pic_size.js"))
    return script_html_file
