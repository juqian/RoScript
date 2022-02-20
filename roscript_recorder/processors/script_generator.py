import os,math
import cv2,numpy
from mako.template import Template
from ruamel import yaml
from roscript_recorder.models.action import SwipeAction, PressKeyboardAction, UnknownAction
from roscript_recorder.constants import *

def export_previews(actions, root_export_dir):
    preview_export_dir = os.path.join(root_export_dir, "previews")
    if os.path.exists(preview_export_dir):
        raise Exception("预览图像文件夹已存在，请先删除原文件夹！")
    os.makedirs(preview_export_dir)
    for action in actions:
        index = action.index
        preview_image = action.get_preview_image()
        preview_name = "preview%d.png"%index
        preview_export_path = os.path.join(preview_export_dir, preview_name)
        cv2.imwrite(preview_export_path, preview_image)

def export_pre_action_frames(actions, root_export_dir):
    frame_export_dir = os.path.join(root_export_dir, "frames")
    if os.path.exists(frame_export_dir):
        raise Exception("设备界面图像文件夹已存在，请先删除原文件夹！")
    os.makedirs(frame_export_dir)
    for action in actions:
        index = action.index
        device_image = action.pre_action_frame.get_device_image()
        frame_name = "frame%d.png"%index
        frame_export_path = os.path.join(frame_export_dir, frame_name)
        cv2.imwrite(frame_export_path, device_image)

def generate_script(actions, root_export_dir):
    if os.path.exists(root_export_dir):
        raise Exception("该脚本文件夹已存在，请先删除原文件夹！")
    os.makedirs(root_export_dir)
    image_export_dir = os.path.join(root_export_dir, "images")
    os.makedirs(image_export_dir)
    script_export_path = os.path.join(root_export_dir, "script.py")
    instructions = []
    for i in range(len(actions)):
        action = actions[i]
        sleep_time = None
        if i > 0:
            fps = action.parent.parameters["fps"]
            pre_action = actions[i-1]
            sleep_frame_count = action.action_fingertip_group[0].index - pre_action.touch_fingertip_groups[-1][-1].index
            sleep_time = math.ceil(sleep_frame_count/fps)
        name = action.get_instruction_name()
        multimedia_parameters = action.get_instruction_parameters()
        string_parameters = []
        image_index = 1
        for parameter in multimedia_parameters:
            if isinstance(parameter, numpy.ndarray):
                widget = parameter
                image_name = "image%d-%d.png"%(action.index, image_index)
                image_export_path = os.path.join(image_export_dir, image_name)
                cv2.imwrite(image_export_path, widget)
                string_parameter = image_name
                image_index += 1
            else:
                string_parameter = parameter
            string_parameters.append(string_parameter)
        combined_parameter = ",".join(['"%s"'%string_parameter for string_parameter in string_parameters])
        instructions.append((name, combined_parameter, sleep_time))
    script_content = Template(SCRIPT_TEMPLATE_CONTENT).render(instructions=instructions, note=None)
    with open(script_export_path, "w") as script_file:
        script_file.write(script_content)
    generate_default_screenshot(actions[0], image_export_dir)

def generate_default_screenshot(action, export_dir):
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)
    screenshot_export_path = os.path.join(export_dir, "default.snapscreen")
    if os.path.exists(screenshot_export_path):
        raise Exception("该screenshot文件已存在，请先删除原文件！")
    device_location = action.parent.parameters["device_location"]
    device_w = device_location.get_width()
    device_h = device_location.get_height()
    screenshot_content = "contour:\n  width: %d\n  height: %d"%(device_w, device_h)
    with open(screenshot_export_path, "w") as f:
        f.write(screenshot_content)

def generate_result_text(actions, export_dir):
    result = {"device":None, "actions":{}}
    device_location = actions[0].parent.parameters["device_location"]
    device_x1, device_y1 = device_location.lt
    device_x2, device_y2 = device_location.rb
    device_location_dict = {"left-top":str((device_x1, device_y1)), "right-bottom":str((device_x2, device_y2))}
    actions_dict = {}
    for i in range(len(actions)):
        action = actions[i]
        action_dict = {}
        action_dict["type"] = action.get_instruction_name()
        if isinstance(action, (SwipeAction, PressKeyboardAction)):
            action_dict["note"] = "|".join(action.get_instruction_parameters())
        elif isinstance(action, UnknownAction):
            pass
        else:
            touch_positions = action.get_key_touch_positions()
            widget_areas = action.get_widget_areas(touch_positions)
            widgets_dict = {}
            for j in range(len(widget_areas)):
                widget_area = widget_areas[j]
                touch_position = touch_positions[j]
                touch_x, touch_y = touch_position
                device_touch_x, device_touch_y = touch_x-device_x1, touch_y-device_y1
                widget_x1, widget_y1 = widget_area.lt
                widget_x2, widget_y2 = widget_area.rb
                device_widget_x1, device_widget_y1 = widget_x1-device_x1, widget_y1-device_y1
                device_widget_x2, device_widget_y2 = widget_x2-device_x1, widget_y2-device_y1
                widget_dict_area = {"left-top":str((device_widget_x1, device_widget_y1)), "right-bottom":str((device_widget_x2, device_widget_y2))}
                widget_dict = {"touch-position":str((device_touch_x, device_touch_y)), "area":widget_dict_area}
                widgets_dict["widget%d"%(j+1)] = widget_dict
            action_dict["widgets"] = widgets_dict
        actions_dict["action%d"%(i+1)] = action_dict
    result = {"device":device_location_dict, "actions":actions_dict}
    result_text_export_path = os.path.join(export_dir, "result.yaml")
    with open(result_text_export_path, "w", encoding="utf-8") as f:
        yaml.dump(result, f, Dumper=yaml.RoundTripDumper,default_flow_style=False,allow_unicode=True, indent=2)

def generate_fingertip_xls(actions, export_dir):
    from openpyxl import Workbook
    
    if "sidelook_touch_threshold" in actions[0].parent.parameters:
        sidelook_touch_threshold = actions[0].parent.parameters["sidelook_touch_threshold"]
    else:
        sidelook_touch_threshold = 0
    frame_total_count = actions[-1].action_fingertip_group[-1].index + 10
    overlook_fingertip_positions = [(0,0) for _ in range(frame_total_count)]
    sidelook_fingertip_positions = [(0,0) for _ in range(frame_total_count)]
    for action in actions:
        for fingertip in action.action_fingertip_group:
            index = fingertip.index
            overlook_fingertip_positions[index] = fingertip.overlook_position if fingertip.overlook_position != None else (0, 0)
            sidelook_fingertip_positions[index] = fingertip.sidelook_position if fingertip.sidelook_position != None else (0, 0)
    
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)
    xls_export_path = os.path.join(export_dir, "fingertip_datas.xlsx")
    wb = Workbook()
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet("Fingertip")
    else:
        ws = wb.get_sheet_by_name(wb.sheetnames[0])
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:C1')
    ws.merge_cells('D1:E1')
    ws.merge_cells('F1:F2')
    ws['A1'] = '序号'
    ws['B1'] = '俯视图指尖'
    ws['D1'] = '侧视图指尖'
    ws['F1'] = '侧视图接触阈值'
    ws['B2'], ws['C2'] = "X", "Y"
    ws['D2'], ws['E2'] = "X", "Y"
    for i in range(frame_total_count):
        overlook_x, overlook_y = overlook_fingertip_positions[i]
        sidelook_x, sidelook_y = sidelook_fingertip_positions[i]
        row = (i, overlook_x, overlook_y, sidelook_x, sidelook_y, sidelook_touch_threshold)
        ws.append(row)
    wb.save(xls_export_path)

def generate(actions, export_dir, dev_mode):
    generate_script(actions, export_dir)
    export_pre_action_frames(actions, export_dir)
    if dev_mode:
        export_previews(actions, export_dir)
        generate_result_text(actions, export_dir)
        generate_fingertip_xls(actions, export_dir)
